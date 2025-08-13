# Date: 2023/01/02
# Warning: this code using tabula-py source code can convert some files that tabula-gui fails to upload and throws the info “Sorry, your file upload could not be processed. Please double-check that the file you uploaded is a valid PDF file and try again”. For example, line 21 in toconvert.csv. I don’t know why!

from wrapper import convert_into
import openpyxl
import os, shutil


# pdf_dir = r'C:\Users\ASUS\Desktop\RA\automation_fail\1012_1062' # Directory of input pdfs
# csv_dir = r'C:\Users\ASUS\Desktop\RA\tabula\tabula-py-results' # Directory of output csvs
# record_xlsx = r'C:\Users\ASUS\Desktop\RA\tabula\failure_list.xlsx' # Path of failure_list.xlsx

pdf_dir = '/Users/melvinleo/Downloads/ATRIUM HOUSE瑧頤'
csv_dir = '/Users/melvinleo/Downloads/HKRE_Test_Folder'
record_xlsx = '/Users/melvinleo/Downloads/HKRE_Test_Folder'


# # This function appends a piece of record to the first column of an excel
# def append_record_to_excel(excel_path, record):
#   # Open the workbook in read-write mode
#   wb = openpyxl.load_workbook(excel_path, read_only=False)
  
#   # Select the sheet you want to append data to
#   sheet = wb['Sheet1']
#   for cell in sheet[1]:
#     # Check if the record is already in the excel
#     if cell.value == record:
#       return
#   # Append data to the next available row, column 1
#   next_row = sheet.max_row + 1
#   sheet.cell(row=next_row, column=1).value = record
#   wb.save(excel_path)


def test_convert_pdf_in_temp(temp_dir, output_dir):
    for file in os.listdir(temp_dir):
        if file.endswith(("PO.PDF", "PR.PDF", "RT.PDF")):
            pdf_path = os.path.join(temp_dir, file)
            csv_filename = os.path.splitext(file)[0] + ".csv"
            csv_path = os.path.join(output_dir, csv_filename)

            print(f"\n🔄 Converting {file} to {csv_filename}...")

            try:
                convert_into(pdf_path, csv_path, 'csv', None, pages='all', guess=True)
                if os.path.exists(csv_path) and os.path.getsize(csv_path) > 0:
                    print(f"✅ Successfully converted {file} → {csv_filename}")
                else:
                    print(f"⚠️ Conversion completed but CSV is empty: {csv_filename}")
            except Exception as e:
                print(f"❌ Failed to convert {file}:", e)



# def append_record_to_excel(excel_path, record):
#     # If the Excel file doesn't exist, create it with Sheet1
#     if not os.path.exists(excel_path):
#         wb = openpyxl.Workbook()
#         sheet = wb.active
#         sheet.title = 'Sheet1'
#         wb.save(excel_path)

#     # Open the workbook
#     wb = openpyxl.load_workbook(excel_path)
#     sheet = wb['Sheet1']

#     # Check if the record already exists in the first row
#     for cell in sheet[1]:
#         if cell.value == record:
#             return

#     # Append data to the next available row, column 1
#     next_row = sheet.max_row + 1
#     sheet.cell(row=next_row, column=1).value = record
#     wb.save(excel_path)


# This function converts all files in pdf_dir outputs to csv_dir
# Records failed files in record_path


def convert(pdf_dir, csv_dir, record_path):
  # Create an empty temporary folder in csv_dir. 
  # If a folder like this already exists, delete and create a new one

  temp_dir = os.path.join(csv_dir, 'temp')
  if os.path.exists(temp_dir):
    shutil.rmtree(temp_dir)
  os.mkdir(temp_dir)
  
  
  # Copy all files in pdf_dir to the temporary folder
  for file in os.listdir(pdf_dir):
    src_path = os.path.join(pdf_dir, file)
    dst_path = os.path.join(temp_dir, file)
    if os.path.isfile(src_path):
      shutil.copy(src_path, dst_path)

  test_convert_pdf_in_temp(temp_dir, csv_dir)


  # Check all csv files in temporary folder
  # If empty, then record it in the excel
  # If with text, then copy the file to csv_dir
  # for file in os.listdir(temp_dir):
  #   if file.endswith('.csv'):
  #     file_size = os.path.getsize(os.path.join(temp_dir, file))
  #     # Check if the file size is 0
  #     if file_size == 0:
  #       append_record_to_excel(record_path, file[:-4])
  #       print(f"{file[:-4]}.pdf converted to empty file using tabula-py.")
  #     else:
  #       shutil.copy(os.path.join(temp_dir, file), os.path.join(csv_dir, file))
  
  # Delete the temporary folder
  shutil.rmtree(temp_dir)

if __name__ == "__main__":
  convert(pdf_dir, csv_dir, record_xlsx)